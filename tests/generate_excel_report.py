"""
GlycoGuard AI - Automated XLSX Test Report & GitHub Step Summary Generator
Reads accumulated test execution results, generates a styled multi-sheet Excel workbook,
and outputs a GitHub Actions Step Summary markdown dashboard.
"""

import os
import json
import time
from datetime import datetime
from pathlib import Path

RESULTS_FILE = Path(__file__).parent / "test_results.json"
EXCEL_OUTPUT = Path(__file__).parent.parent / "GlycoGuard_CI_CD_Test_Report.xlsx"
SUMMARY_MD_OUTPUT = Path(__file__).parent.parent / "test_summary.md"


def load_results():
    """Load records from test_results.json."""
    if not RESULTS_FILE.exists():
        print(f"[REPORT WARNING] {RESULTS_FILE} not found. Creating empty report.")
        return []
    try:
        with open(RESULTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[REPORT ERROR] Could not read test_results.json: {e}")
        return []


def generate_excel(records):
    """Build a styled, professional Excel workbook with Executive Summary & Detailed Logs."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[REPORT ERROR] openpyxl not installed. Skipping Excel generation.")
        return

    wb = openpyxl.Workbook()

    # Color Palette
    HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Navy
    HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
    SUBTITLE_FONT = Font(name="Segoe UI", size=11, italic=True, color="475569")
    BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
    NORMAL_FONT = Font(name="Segoe UI", size=10)
    
    PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Mint Green
    PASS_FONT = Font(name="Segoe UI", size=10, bold=True, color="065F46")
    
    FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
    FAIL_FONT = Font(name="Segoe UI", size=10, bold=True, color="7F1D1D")
    
    SKIP_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Amber
    SKIP_FONT = Font(name="Segoe UI", size=10, bold=True, color="92400E")

    THIN_BORDER = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    # -------------------------------------------------------------
    # SHEET 1: Executive Summary
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary["A2"] = "🩺 GlycoGuard AI - CI/CD Automated Test Execution Report"
    ws_summary["A2"].font = TITLE_FONT
    ws_summary["A3"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')} | Environment: GitHub Actions CI/CD"
    ws_summary["A3"].font = SUBTITLE_FONT

    # Calculate Metrics
    total_tests = len(records)
    passed_tests = sum(1 for r in records if r.get("status") == "PASS")
    failed_tests = sum(1 for r in records if r.get("status") == "FAIL")
    skipped_tests = sum(1 for r in records if r.get("status") == "SKIPPED")
    pass_percentage = round((passed_tests / total_tests * 100), 1) if total_tests > 0 else 0.0
    total_duration = round(sum(r.get("execution_time", 0.0) for r in records), 2)

    # Status Banner Card
    overall_status = "ALL TESTS PASSED ✅" if failed_tests == 0 and total_tests > 0 else "FAILURES DETECTED ❌"
    status_fill = PASS_FILL if failed_tests == 0 and total_tests > 0 else FAIL_FILL
    status_font = PASS_FONT if failed_tests == 0 and total_tests > 0 else FAIL_FONT

    ws_summary["A5"] = "Overall Verdict"
    ws_summary["B5"] = overall_status
    ws_summary["A5"].font = BOLD_FONT
    ws_summary["B5"].font = status_font
    ws_summary["B5"].fill = status_fill

    # Execution Metadata Table
    meta = [
        ("Commit SHA", os.getenv("GITHUB_SHA", "Local-Dev-Run")[:10]),
        ("Branch", os.getenv("GITHUB_REF_NAME", "main")),
        ("Frontend URL", os.getenv("FRONTEND_URL", "https://lakshmiankal.github.io/GlycoGuard-AI/")),
        ("Backend URL", os.getenv("BACKEND_URL", "https://glycoguard-api.onrender.com")),
        ("Total Execution Time", f"{total_duration} seconds"),
        ("Pass Rate", f"{pass_percentage}%")
    ]

    for idx, (k, v) in enumerate(meta, start=6):
        ws_summary[f"A{idx}"] = k
        ws_summary[f"B{idx}"] = v
        ws_summary[f"A{idx}"].font = BOLD_FONT
        ws_summary[f"B{idx}"].font = NORMAL_FONT
        ws_summary[f"A{idx}"].border = THIN_BORDER
        ws_summary[f"B{idx}"].border = THIN_BORDER

    # Summary by Category Table
    categories = ["Unit Test", "API Test", "Selenium E2E Test", "Security/Vulnerability Test"]
    cat_start = 14
    ws_summary[f"A{cat_start}"] = "Test Category"
    ws_summary[f"B{cat_start}"] = "Total"
    ws_summary[f"C{cat_start}"] = "Passed"
    ws_summary[f"D{cat_start}"] = "Failed"
    ws_summary[f"E{cat_start}"] = "Skipped"
    ws_summary[f"F{cat_start}"] = "Pass Rate"
    ws_summary[f"G{cat_start}"] = "Status"

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        cell = ws_summary[f"{col}{cat_start}"]
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if col != "A" else "left")

    for i, cat in enumerate(categories, start=cat_start + 1):
        c_records = [r for r in records if r.get("test_type") == cat]
        c_tot = len(c_records)
        c_pass = sum(1 for r in c_records if r.get("status") == "PASS")
        c_fail = sum(1 for r in c_records if r.get("status") == "FAIL")
        c_skip = sum(1 for r in c_records if r.get("status") == "SKIPPED")
        c_rate = f"{round(c_pass / c_tot * 100, 1)}%" if c_tot > 0 else "N/A"
        c_stat = "PASS" if c_fail == 0 and c_tot > 0 else ("FAIL" if c_fail > 0 else "NO TESTS")

        ws_summary[f"A{i}"] = cat
        ws_summary[f"B{i}"] = c_tot
        ws_summary[f"C{i}"] = c_pass
        ws_summary[f"D{i}"] = c_fail
        ws_summary[f"E{i}"] = c_skip
        ws_summary[f"F{i}"] = c_rate
        ws_summary[f"G{i}"] = c_stat

        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            cell = ws_summary[f"{col}{i}"]
            cell.font = BOLD_FONT if col in ["A", "G"] else NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col != "A" else "left")
            if col == "G":
                cell.fill = PASS_FILL if c_stat == "PASS" else (FAIL_FILL if c_stat == "FAIL" else SKIP_FILL)
                cell.font = PASS_FONT if c_stat == "PASS" else (FAIL_FONT if c_stat == "FAIL" else SKIP_FONT)

    # -------------------------------------------------------------
    # SHEET 2: Detailed Test Results
    # -------------------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Test Results")
    ws_details.views.sheetView[0].showGridLines = True

    headers = [
        "Test Case ID",
        "Test Type",
        "Test Description",
        "Expected Result",
        "Actual Result",
        "Status",
        "Execution Time (s)",
        "Failure Details"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, r in enumerate(records, start=2):
        status = r.get("status", "UNKNOWN")
        
        ws_details.cell(row=row_idx, column=1, value=r.get("test_id", "")).alignment = Alignment(horizontal="center")
        ws_details.cell(row=row_idx, column=2, value=r.get("test_type", "")).alignment = Alignment(horizontal="center")
        ws_details.cell(row=row_idx, column=3, value=r.get("test_description", ""))
        ws_details.cell(row=row_idx, column=4, value=r.get("expected_result", ""))
        ws_details.cell(row=row_idx, column=5, value=r.get("actual_result", ""))
        
        status_cell = ws_details.cell(row=row_idx, column=6, value=status)
        status_cell.alignment = Alignment(horizontal="center")
        if status == "PASS":
            status_cell.fill = PASS_FILL
            status_cell.font = PASS_FONT
        elif status == "FAIL":
            status_cell.fill = FAIL_FILL
            status_cell.font = FAIL_FONT
        else:
            status_cell.fill = SKIP_FILL
            status_cell.font = SKIP_FONT

        ws_details.cell(row=row_idx, column=7, value=r.get("execution_time", 0.0)).alignment = Alignment(horizontal="center")
        ws_details.cell(row=row_idx, column=8, value=r.get("failure_details", "N/A"))

        for col_idx in range(1, 9):
            c = ws_details.cell(row=row_idx, column=col_idx)
            c.border = THIN_BORDER
            if col_idx != 6:
                c.font = NORMAL_FONT

    # Auto-adjust column widths for readability
    for ws in [ws_summary, ws_details]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if "\n" in val:
                    val = max(val.split("\n"), key=len)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    wb.save(str(EXCEL_OUTPUT))
    print(f"[SUCCESS] Excel Test Report generated: {EXCEL_OUTPUT}")


def generate_step_summary(records):
    """Write rich GitHub-flavored Markdown step summary."""
    total = len(records)
    passed = sum(1 for r in records if r.get("status") == "PASS")
    failed = sum(1 for r in records if r.get("status") == "FAIL")
    skipped = sum(1 for r in records if r.get("status") == "SKIPPED")
    pass_pct = round((passed / total * 100), 1) if total > 0 else 0.0
    duration = round(sum(r.get("execution_time", 0.0) for r in records), 2)
    commit = os.getenv("GITHUB_SHA", "Manual-Local-Run")[:8]
    branch = os.getenv("GITHUB_REF_NAME", "main")

    verdict_badge = "✅ **PASSED (ALL CHECKS GREEN)**" if failed == 0 and total > 0 else "❌ **FAILED (ISSUES DETECTED)**"
    verdict_plain = "✅ PASSED" if failed == 0 and total > 0 else "❌ FAILED"

    md = f"""# 🩺 GlycoGuard AI - CI/CD Automated Test Report

### 🎯 Overall Status: {verdict_badge}

| Metric | Value |
| :--- | :--- |
| **Commit SHA** | `{commit}` |
| **Branch** | `{branch}` |
| **Total Test Duration** | **{duration}s** |
| **Pass Rate** | **{pass_pct}%** |
| **Total Tests Executed** | **{total}** (Passed: {passed}, Failed: {failed}, Skipped: {skipped}) |

---

### 📊 Results Breakdown by Test Category

| Test Category | Total | Passed | Failed | Skipped | Pass Rate | Status |
| :--- | :---: | :---: | :---: | :---: | :---: | :---: |
"""

    categories = [
        ("Unit Tests", "Unit Test"),
        ("Live API Tests (Render)", "API Test"),
        ("Selenium E2E Tests (GitHub Pages)", "Selenium E2E Test"),
        ("Security & Vulnerability", "Security/Vulnerability Test")
    ]

    for label, cat_name in categories:
        c_records = [r for r in records if r.get("test_type") == cat_name]
        c_tot = len(c_records)
        c_pass = sum(1 for r in c_records if r.get("status") == "PASS")
        c_fail = sum(1 for r in c_records if r.get("status") == "FAIL")
        c_skip = sum(1 for r in c_records if r.get("status") == "SKIPPED")
        c_rate = f"{round(c_pass / c_tot * 100, 1)}%" if c_tot > 0 else "N/A"
        c_badge = "✅ PASS" if c_fail == 0 and c_tot > 0 else ("❌ FAIL" if c_fail > 0 else "⚠️ NONE")
        md += f"| **{label}** | {c_tot} | {c_pass} | {c_fail} | {c_skip} | {c_rate} | {c_badge} |\n"

    md += f"| **TOTAL** | **{total}** | **{passed}** | **{failed}** | **{skipped}** | **{pass_pct}%** | {verdict_badge} |\n\n"

    # Failed Tests Section
    failed_items = [r for r in records if r.get("status") == "FAIL"]
    if failed_items:
        md += "### ❌ Failed Tests Breakdown\n\n"
        md += "| Test ID | Test Category | Description | Failure Details |\n"
        md += "| :--- | :--- | :--- | :--- |\n"
        for item in failed_items:
            clean_err = item.get("actual_result", "").replace("\n", " ").replace("|", "/")[:120]
            md += f"| `{item.get('test_id')}` | {item.get('test_type')} | {item.get('test_description')} | `{clean_err}` |\n"
        md += "\n"
    else:
        md += "### 🎉 Test Execution Verdict\n"
        md += "> All automated test suites passed successfully with zero failures! System is ready for production.\n\n"

    md += "📥 **Artifacts Generated**: `GlycoGuard_CI_CD_Test_Report.xlsx` is uploaded as an artifact in this workflow run.\n"

    # Save to local summary markdown
    with open(SUMMARY_MD_OUTPUT, "w", encoding="utf-8") as f:
        f.write(md)

    # Write to GitHub Step Summary if running in GitHub Actions
    gh_summary_file = os.getenv("GITHUB_STEP_SUMMARY")
    if gh_summary_file:
        try:
            with open(gh_summary_file, "a", encoding="utf-8") as f:
                f.write(md)
            print("[SUCCESS] Appended test summary to $GITHUB_STEP_SUMMARY.")
        except Exception as e:
            print(f"[WARNING] Could not write to $GITHUB_STEP_SUMMARY: {e}")


def main():
    records = load_results()
    print(f"--- Loaded {len(records)} test execution records ---")
    generate_excel(records)
    generate_step_summary(records)


if __name__ == "__main__":
    main()
